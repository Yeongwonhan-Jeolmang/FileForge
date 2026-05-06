"""
file_info.py — Read every property we can about a file.
Returns a rich FileInfo dataclass used across the app.
"""

from __future__ import annotations
import os
import stat
import time
import platform
import hashlib
import mimetypes
from dataclasses import dataclass, field
from typing import Optional
from pathlib import Path

# ── Optional deps (graceful degradation) ──────────────────────────────────
try:
    from mutagen._file import File as MutagenFile
    HAS_MUTAGEN = True
except ImportError:
    HAS_MUTAGEN = False

try:
    from PIL import Image
    HAS_PIL = True
except ImportError:
    HAS_PIL = False
try:
    from PyPDF2 import PdfReader
    HAS_PYPDF2 = True
except ImportError:
    HAS_PYPDF2 = False

try:
    from docx import Document as DocxDocument
    HAS_DOCX = True
except ImportError:
    HAS_DOCX = False

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    from pptx import Presentation
    HAS_PPTX = True
except ImportError:
    HAS_PPTX = False

try:
    from pymediainfo import MediaInfo
    HAS_MEDIainfo = True
except ImportError:
    HAS_MEDIainfo = False
# ── Data structures ────────────────────────────────────────────────────────

@dataclass
class PermissionInfo:
    owner_read: bool
    owner_write: bool
    owner_exec: bool
    group_read: bool
    group_write: bool
    group_exec: bool
    other_read: bool
    other_write: bool
    other_exec: bool
    octal: str
    symbolic: str
 
 
@dataclass
class ImageMetadata:
    width: int
    height: int
    mode: str
    format: str
    dpi: Optional[tuple]
    has_exif: bool
    exif: dict = field(default_factory=dict)
 
 
@dataclass
class AudioMetadata:
    duration: Optional[float]
    bitrate: Optional[int]
    sample_rate: Optional[int]
    channels: Optional[int]
    tags: dict = field(default_factory=dict)
 
 
@dataclass
class DocumentMetadata:
    doc_type: str = ''
    title: str = ''
    author: str = ''
    producer: str = ''
    subject: str = ''
    pages: Optional[int] = None
    software: str = ''
    topic: str = ''
    comment: str = ''
    snippet: str = ''
 
 
@dataclass
class VideoMetadata:
    duration: Optional[float] = None
    width: Optional[int] = None
    height: Optional[int] = None
    bitrate: Optional[int] = None
    codec: str = ''
    format: str = ''
 
 
@dataclass
class FileInfo:
    # Basics
    path: str
    name: str
    stem: str
    suffix: str
    size: int
    size_human: str
 
    # Timestamps
    created: float
    modified: float
    accessed: float
    created_str: str
    modified_str: str
    accessed_str: str
 
    # Type
    mime_type: str
    kind: str          # "image", "audio", "video", "document", "archive", "text", "other"
    is_symlink: bool
    is_hidden: bool
 
    # Permissions
    permissions: PermissionInfo
    owner: str
    group: str
 
    # Hashes (computed lazily)
    md5: Optional[str] = None
    sha1: Optional[str] = None
    sha256: Optional[str] = None
    sha512: Optional[str] = None
    blake2: Optional[str] = None
 
    # Rich metadata
    image_meta: Optional[ImageMetadata] = None
    audio_meta: Optional[AudioMetadata] = None
    document_meta: Optional[DocumentMetadata] = None
    video_meta: Optional[VideoMetadata] = None
 
    # Extended attributes and platform-specific extras
    xattrs: dict = field(default_factory=dict)
    windows_ads: dict = field(default_factory=dict)

# ── Helpers ────────────────────────────────────────────────────────────────

_SIZE_UNITS = ["B", "KB", "MB", "GB", "TB", "PB"]

def human_size(n: int) -> str:
    val = float(n)
    for unit in _SIZE_UNITS:
        if val < 1024:
            return f"{val:.2f}" if unit != "B" else f"{int(val)} B"
        val /= 1024
    return f"{val:.2f} PB"

def fmt_time(ts: float) -> str:
    return time.strftime("%Y-%m-%d  %H:%M:%S", time.localtime(ts))

def _parse_permissions(mode: int) -> PermissionInfo:
    s = stat.filemode(mode)
    return PermissionInfo(
        owner_read  = bool(mode & stat.S_IRUSR),
        owner_write = bool(mode & stat.S_IWUSR),
        owner_exec  = bool(mode & stat.S_IXUSR),
        group_read  = bool(mode & stat.S_IRGRP),
        group_write = bool(mode & stat.S_IWGRP),
        group_exec  = bool(mode & stat.S_IXGRP),
        other_read  = bool(mode & stat.S_IROTH),
        other_write = bool(mode & stat.S_IWOTH),
        other_exec  = bool(mode & stat.S_IXOTH),
        octal       = oct(stat.S_IMODE(mode)),
        symbolic    = s,
    )

def _owner_group(st) -> tuple[str, str]:
    if platform.system() == "Windows":
        return "N/A", "N/A"
    try:
        import pwd, grp
        owner = pwd.getpwuid(st.st_uid).pw_name
        group = grp.getgrgid(st.st_gid).gr_name
        return owner, group
    except Exception:
        return str(getattr(st, "st_uid", "?")), str(getattr(st, "st_gid", "?"))

def _classify(mime: str, suffix: str) -> str:
    if mime.startswith("image/"):         return "image"
    if mime.startswith("audio/"):         return "audio"
    if mime.startswith("video/"):         return "video"
    if mime.startswith("text/"):          return "text"
    if "pdf" in mime or suffix in (".doc", ".docx", ".odt", ".xls", ".xlsx", ".ppt", ".pptx"):
        return "document"
    if suffix in (".zip", ".tar", ".gz", ".bz2", ".xz", ".7z", ".rar", ".zst"):
        return "archive"
    return "other"

def _image_meta(path: str) -> Optional[ImageMetadata]:
    if not HAS_PIL:
        return None
    try:
        with Image.open(path) as img:
            exif_raw = {}
            has_exif = False
            if hasattr(img, "_getexif") and img._getexif():
                has_exif = True
                raw = img._getexif() or {}
                from PIL.ExifTags import TAGS
                for tag_id, val in raw.items():
                    tag = TAGS.get(tag_id, tag_id)
                    try:
                        exif_raw[str(tag)] = str(val)
                    except Exception:
                        pass
            return ImageMetadata(
                width=img.width,
                height=img.height,
                mode=img.mode,
                format=img.format or "?",
                dpi=img.info.get("dpi"),
                has_exif=has_exif,
                exif=exif_raw,
            )
    except Exception:
        return None

def _audio_meta(path: str) -> Optional[AudioMetadata]:
    if not HAS_MUTAGEN:
        return None
    try:
        mf = MutagenFile(path, easy=True)
        if mf is None:
            return None
        info = getattr(mf, "info", None)
        tags = {}
        if mf.tags:
            for k, v in mf.tags.items():
                tags[k] = str(v[0]) if isinstance(v, list) else str(v)
        return AudioMetadata(
            duration    = getattr(info, "length", None),
            bitrate     = getattr(info, "bitrate", None),
            sample_rate = getattr(info, "sample_rate", None),
            channels    = getattr(info, "channels", None),
            tags        = tags,
        )
    except Exception:
        return None


def _document_meta(path: str, suffix: str) -> Optional[DocumentMetadata]:
    if suffix == '.pdf' and HAS_PYPDF2:
        try:
            reader = PdfReader(path)
            meta = reader.metadata or {}
            title = getattr(meta, 'title', '') or ''
            author = getattr(meta, 'author', '') or ''
            producer = getattr(meta, 'producer', '') or ''
            subject = getattr(meta, 'subject', '') or ''
            software = getattr(meta, 'producer', '') or ''
            pages = len(reader.pages)
            snippet = ''
            if pages:
                try:
                    snippet = reader.pages[0].extract_text() or ''
                    snippet = snippet[:1200].replace('\n', ' ')
                except Exception:
                    snippet = ''
            return DocumentMetadata(
                doc_type='PDF', title=title, author=author,
                producer=producer, subject=subject,
                pages=pages, software=software, comment=str(getattr(meta, 'keywords', '') or ''),
                topic=subject, snippet=snippet,
            )
        except Exception:
            return None

    if suffix == '.docx' and HAS_DOCX:
        try:
            doc = DocxDocument(path)
            props = doc.core_properties
            return DocumentMetadata(
                doc_type='DOCX', title=props.title or '', author=props.author or '',
                subject=props.subject or '', software=props.last_modified_by or '',
                comment=props.comments or '', snippet='',
            )
        except Exception:
            return None

    if suffix in ('.xlsx', '.xls') and HAS_OPENPYXL:
        try:
            from openpyxl import load_workbook
            wb = load_workbook(path, read_only=True, data_only=True)
            return DocumentMetadata(
                doc_type='Spreadsheet', title=Path(path).name,
                pages=len(wb.sheetnames), snippet='',
            )
        except Exception:
            return None

    if suffix == '.pptx' and HAS_PPTX:
        try:
            prs = Presentation(path)
            return DocumentMetadata(
                doc_type='Presentation', title=Path(path).name,
                pages=len(prs.slides), snippet='',
            )
        except Exception:
            return None

    return None


def _video_meta(path: str) -> Optional[VideoMetadata]:
    if not HAS_MEDIainfo:
        return None
    try:
        info = MediaInfo.parse(path)
        for track in info.tracks:
            if track.track_type == 'Video':
                duration_ms = getattr(track, 'duration', None)
                duration = duration_ms / 1000 if duration_ms else None
                return VideoMetadata(
                    duration=duration,
                    width=getattr(track, 'width', None),
                    height=getattr(track, 'height', None),
                    bitrate=getattr(track, 'bit_rate', None) or getattr(track, 'bitrate', None),
                    codec=getattr(track, 'codec', '') or getattr(track, 'format', ''),
                    format=getattr(track, 'format', ''),
                )
    except Exception:
        pass
    return None


def _windows_ads(path: str) -> dict:
    if platform.system() != 'Windows':
        return {}
    try:
        import subprocess
        output = subprocess.check_output(['cmd.exe', '/c', 'dir', '/r', str(Path(path))], stderr=subprocess.DEVNULL, text=True)
        ads = {}
        for line in output.splitlines():
            if ':$DATA' in line and 'Directory of' not in line and line.strip():
                parts = line.split()
                if parts and parts[0].isdigit():
                    size = int(parts[0])
                    stream = parts[-1]
                    ads[stream] = size
        return ads
    except Exception:
        return {}


def _xattrs(path: str) -> dict:
    if not hasattr(os, "getxattr"):
        return {}
    try:
        attrs = {}
        for name in os.listxattr(path):
            try:
                attrs[name] = os.getxattr(path, name).decode(errors="replace")
            except Exception:
                attrs[name] = "<unreadable>"
        return attrs
    except Exception:
        return {}
 
 
# ── Public API ─────────────────────────────────────────────────────────────
 
def read_file_info(path: str) -> FileInfo:
    """Return a fully populated FileInfo for the given path."""
    p = Path(path)
    st = p.lstat()
    mode = st.st_mode
 
    mime, _ = mimetypes.guess_type(str(p))
    mime = mime or "application/octet-stream"
    suffix = p.suffix.lower()
    kind = _classify(mime, suffix)
 
    is_sym = p.is_symlink()
    is_hidden = p.name.startswith(".") or (
        platform.system() == "Windows" and bool(st.st_file_attributes & 2) if hasattr(st, "st_file_attributes") else False
    )
 
    created_ts = getattr(st, "st_birthtime", st.st_ctime)
    owner, group = _owner_group(st)
 
    info = FileInfo(
        path         = str(p),
        name         = p.name,
        stem         = p.stem,
        suffix       = p.suffix,
        size         = st.st_size,
        size_human   = human_size(st.st_size),
        created      = created_ts,
        modified     = st.st_mtime,
        accessed     = st.st_atime,
        created_str  = fmt_time(created_ts),
        modified_str = fmt_time(st.st_mtime),
        accessed_str = fmt_time(st.st_atime),
        mime_type    = mime,
        kind         = kind,
        is_symlink   = is_sym,
        is_hidden    = is_hidden,
        permissions  = _parse_permissions(mode),
        owner        = owner,
        group        = group,
        image_meta    = _image_meta(path) if kind == "image" else None,
        audio_meta    = _audio_meta(path) if kind == "audio" else None,
        document_meta = _document_meta(path, suffix) if kind == "document" else None,
        video_meta    = _video_meta(path) if kind == "video" else None,
        xattrs        = _xattrs(path),
        windows_ads   = _windows_ads(path),
    )
    return info
 
 
def compute_hashes(path: str, algorithms: list[str] = None, progress_cb=None) -> dict[str, str]:
    """Compute file hashes. algorithms: list of 'md5','sha1','sha256','sha512','blake2'."""
    if algorithms is None:
        algorithms = ["md5", "sha1", "sha256"]

    # Handle blake2 specially as it needs a constructor
    hashers = {}
    for a in algorithms:
        if a == "blake2":
            hashers[a] = hashlib.blake2b()
        else:
            hashers[a] = hashlib.new(a)

    total = os.path.getsize(path)
    done = 0
    chunk = 65536
 
    with open(path, "rb") as f:
        while True:
            data = f.read(chunk)
            if not data:
                break
            for h in hashers.values():
                h.update(data)
            done += len(data)
            if progress_cb:
                progress_cb(done, total)
 
    return {a: hashers[a].hexdigest() for a in algorithms}